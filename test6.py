import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pygame

def create_pivot_table_summary(file_path):
    """
    在Excel文件中创建一个新的sheet页，包含学生信息的透视表汇总
    
    参数:
        file_path (str): Excel文件路径
    """
    # 读取Excel文件
    df = pd.read_excel(file_path, sheet_name='student_data')
    
    # 使用pandas创建透视表
    # 这里假设你想按性别和班级统计学生人数，你可以根据需要修改这些字段
    pivot_df = df.pivot_table(
        index=['班级', '性别'],  # 行字段
        values=['姓名'],        # 值字段
        aggfunc='count',        # 计数
        fill_value=0            # 填充空值为0
    )
    
    # 重置索引，使透视表结果更易读
    pivot_df = pivot_df.reset_index()
    
    # 加载Excel工作簿
    wb = load_workbook(file_path)
    
    # 创建新的sheet页
    ws = wb.create_sheet(title="学生信息汇总")
    
    # 将透视表数据写入新的sheet页
    for r in dataframe_to_rows(pivot_df, index=False, header=True):
        ws.append(r)
    
    # 保存工作簿
    wb.save(file_path)
    print(f"已成功在 {file_path} 中创建学生信息汇总sheet页")

# 使用示例
file_path = "student_data.xlsx"  # 替换为你的Excel文件路径
create_pivot_table_summary(file_path)
